from openpyxl import load_workbook  # 讀取 Excel 文件
import urllib.request  # 更新資料用的

#更新 data.xlsx函數
#試算表:https://docs.google.com/spreadsheets/d/1svRcTy-waZP1bpnVj0K3jTsIrUHeDLVMHR1Wk47z8To/edit?gid=0#gid=0
def load():
    try:
        url = "https://docs.google.com/spreadsheets/d/1svRcTy-waZP1bpnVj0K3jTsIrUHeDLVMHR1Wk47z8To/export?format=xlsx&gid=0"
        urllib.request.urlretrieve(url, "data.xlsx")
    except:
        pass

#有更新檔案再打開
#load()
# 讀取 Excel 文件

wb = load_workbook("data.xlsx",read_only=True)
sheet = wb.active

# 將資料讀取到 chemicals 中給 main.py 使用
chemicals = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    Name, ion, H, OH = row
    chemicals[Name] = {"ion": ion, "H+": H, "OH-": OH, "H+_need": OH}

# debug: 輸出確認資料是否正確讀取
#for key in chemicals:
#    print(key, chemicals[key])


    
